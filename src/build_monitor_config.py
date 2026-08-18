#!/usr/bin/env python3
"""
Parse the monitoring spreadsheet (config/links.xlsx) into config/monitor_sites.json.

Every named row becomes a monitored site, EXCEPT the entries in EXCLUDED_NAMES
(法规 / 专利及文献), which are still under discussion and need a separate
whole-web search mechanism, so they are deliberately left out for now.

Rows that are neither excluded nor carry a valid http(s) news URL are reported
as "needs_url" (a warning), so nothing is ever silently dropped.

Re-run this whenever links.xlsx changes:
    python3 src/build_monitor_config.py
"""

import json
import os
import re
import sys

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(BASE_DIR, "config", "links.xlsx")
OUT_PATH = os.path.join(BASE_DIR, "config", "monitor_sites.json")
SHEET_NAME = "监测名单"

# Rows deliberately left out of monitoring for now (still under discussion —
# they need a whole-web search mechanism rather than a single news page).
EXCLUDED_NAMES = {
    "法规(GRAS,EFSA,FEMA,澳新等多国家)",
    "专利及文献",
}

# Sites kept in the config but NOT monitored (enabled=False). They have no
# usable news page (only a homepage), so scraping them just produces noise.
# Listed here rather than dropped so they're visible and easy to re-enable if
# a real news page turns up later.
DISABLED_NAMES = {
    "HHOYA",
    "NICKS",
}

# Per-site article-path overrides. The scraper's is_probable_article() keeps a
# link only if it looks like an article (in the news dir / has a /YYYY/ segment
# / path contains "news"). If a site's real articles live under a different path
# and get wrongly filtered out, add that path fragment here to whitelist it.
# Example: "Tate & Lyle": ["/articles/"].
ALLOW_PATHS = {
    # "站点名": ["/path-fragment/"],
    # 浩天: real articles are /new-150-160.html … /new-152-158.html ("new-",
    # NOT "news-"). The list/pagination links are /news-6-N.html. Articles have
    # only 2 hyphens and no "news" substring, so the generic rules (which need a
    # 3-hyphen slug or the "news" token) drop them — whitelist the "/new-" prefix.
    # "/new-" does not match "/news-6-2.html" (that's "news" + "-"), so pagination
    # stays filtered. Ground truth: real-driver diag on 2026-07-27 (12 articles
    # kept, all nav/product/pagination dropped).
    "浩天": ["/new-"],
    # Morita Kagaku Kogyo: real articles are /en/news/news1.html … news12.html.
    # The filename "newsN" has no hyphen and isn't a bare section slug, so the
    # generic rules drop it. Whitelisting "/news/news" keeps the 12 articles while
    # /en/{about,products,company,privacy,…}.html nav stays filtered. Ground
    # truth: real-driver diag on 2026-07-27.
    "Morita Kagaku Kogyo": ["/news/news"],
}

# Per-site "articles must be dated" switch — the OPPOSITE of ALLOW_PATHS (a
# tightening, not a loosening). For sites whose real articles ALL carry a /YYYY/
# path segment while their section-index pages don't, this drops the index pages
# that the generic rules would otherwise keep (hyphen-slug .html under /news).
# dsm-firmenich: real articles are /news/<section>/2026/<slug>.html, but the list
# page also links bare indexes (/news/share-buy-back.html, investors/…/articles-
# charters.html) whose "body" is aggregate/nav text, not an article. Ground truth:
# real-driver diag on 2026-07-28 (diag_dsm.py) — all 6 dated 2026 articles kept,
# both undated index pages dropped. Whitelisted in ALLOW_PATHS still wins over it.
REQUIRE_YEAR = {
    "dsm-firmenich",
}

# Per-site "English headlines only" switch. For sites that mix in non-English
# TRANSLATIONS of the same story with no /es/ /pt/ path or lang metadata to catch
# them (all served as <html lang="en-US">), so R4's path-mirror rule misses them
# and the title text is the only signal. IFF's newsroom lists Spanish/Portuguese
# versions of the same story alongside the English ones. Conservative detection
# (Romance accent + >=2 Romance stopwords) never drops a real English headline.
# Ground truth: real-driver diag on 2026-07-28.
ENGLISH_ONLY = {
    "IFF",
}

# Per-site content group. Default (unlisted) is "competitor" — commercial
# competitor/industry news summarized with the competitive-intelligence prompt.
# "regulatory" marks government / legal / patent / standards sources (卫健委,
# and future GRAS/EFSA/FEMA-style additions): their articles are official
# notices, so run_monitor routes them to the regulatory-affairs prompt instead.
GROUPS = {
    "国家卫健委": "regulatory",
}

# Per-site JSON-API scraping config. A few sites (Vue/SPA) render their news
# list and bodies from a backend JSON feed rather than server HTML, so the
# browser-based extractor sees an empty page. For those, fetch_monitor_sites'
# extract_items_api() reads the feed directly using this descriptor. Keyed by
# site name; merged verbatim into the site dict (adds "fetch_mode": "json_api"
# plus the "api" block). Kept HERE (not just in the JSON) so a rebuild from
# links.xlsx doesn't silently drop it — the "重建覆盖" lesson.
JSON_API = {
    "元气森林": {
        "fetch_mode": "json_api",
        "api": {
            "list_url": "https://www.yuanqisenlin.com/web/api/newsList",
            "list_items_path": "data.data",
            "list_title_key": "title",
            "list_date_key": "day",
            "list_id_key": "id",
            "list_summary_key": "short_desc",
            "detail_url": "https://www.yuanqisenlin.com/web/api/newsDetail",
            "detail_id_param": "id",
            "detail_content_path": "data.content",
            "article_url_template": "https://www.yuanqisenlin.com/newsInfo?id={id}",
        },
    },
}


def split_keywords(raw):
    """Split a keyword cell into a clean list. Separators: ; , 、 and newlines."""
    if not raw:
        return []
    parts = re.split(r"[;,、\n]+", str(raw))
    seen = set()
    out = []
    for p in parts:
        kw = p.strip()
        # Drop empties and trailing notes like "公司动态新闻" that are descriptions,
        # not actual keywords. We keep them anyway (harmless), just deduped.
        if kw and kw.lower() not in seen:
            seen.add(kw.lower())
            out.append(kw)
    return out


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"❌ Spreadsheet not found: {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    sites = []
    excluded = []   # intentionally left out (法规 / 专利及文献)
    needs_url = []  # not excluded but still missing a valid http(s) news URL
    for r in rows[1:]:  # skip header
        name = (str(r[0]).strip() if r[0] else "")
        keywords_raw = (str(r[1]).strip() if len(r) > 1 and r[1] else "")
        homepage = (str(r[2]).strip() if len(r) > 2 and r[2] else "")
        news_url = (str(r[3]).strip() if len(r) > 3 and r[3] else "")

        if not name:
            continue

        # Deliberately excluded rows (still under discussion).
        if name in EXCLUDED_NAMES:
            excluded.append({"name": name, "reason": "under discussion (needs web search)"})
            continue

        # Every other row is meant to be monitored. If it lacks a usable URL,
        # flag it loudly instead of silently dropping it.
        if not news_url.lower().startswith("http"):
            needs_url.append({"name": name, "value": news_url or "(empty)"})
            continue

        keywords = split_keywords(keywords_raw)
        # Always include the company/platform name as a keyword so its own
        # releases match even when the keyword cell only lists product names.
        if name and name.lower() not in [k.lower() for k in keywords]:
            keywords.append(name)
        site = {
            "name": name,
            "keywords": keywords,
            "homepage": homepage,
            "news_url": news_url,
            "enabled": name not in DISABLED_NAMES,
        }
        if name in ALLOW_PATHS:
            site["allow_paths"] = ALLOW_PATHS[name]
        if name in REQUIRE_YEAR:
            site["require_year"] = True
        if name in ENGLISH_ONLY:
            site["english_only"] = True
        if name in JSON_API:
            site.update(JSON_API[name])
        if name in GROUPS:
            site["group"] = GROUPS[name]
        sites.append(site)

    config = {
        "sites": sites,
        "excluded": excluded,
        "needs_url": needs_url,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"✅ Wrote {OUT_PATH}")
    print(f"   Monitored sites: {len(sites)}")
    print(f"   Excluded (under discussion): {len(excluded)}")
    if needs_url:
        print(f"   ⚠️  Rows still missing a valid news URL: {len(needs_url)}")
        for item in needs_url:
            print(f"       - {item['name']}  (current value: {item['value']})")


if __name__ == "__main__":
    main()
