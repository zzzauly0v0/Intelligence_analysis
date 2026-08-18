"""Turn the monitoring spreadsheet into config/monitor_sites.json.

Every named row becomes a monitored site except the ones in ``EXCLUDED_NAMES``,
which need a whole-web search mechanism rather than a fixed news page. A row that
is neither excluded nor carries an http(s) news URL is reported as "needs_url", so
nothing is dropped silently.

Re-run whenever links.xlsx changes:
    python -m app.services.crawler.config_builder
"""

from __future__ import annotations

import json
import logging
import re
import sys
from collections.abc import Iterable, Sequence
from typing import Any

from app.services.crawler.common.paths import LINKS_XLSX, SITES_PATH
from app.services.crawler.delivery.reporting import configure_console

logger = logging.getLogger(__name__)

SHEET_NAME = "监测名单"

EXCLUDED_NAMES = frozenset(
    {
        "法规(GRAS,EFSA,FEMA,澳新等多国家)",
        "专利及文献",
    }
)

DISABLED_NAMES = frozenset({"HHOYA", "NICKS"})

# Per-site scraping overrides. See urls.py for what each one relaxes or tightens.
ALLOW_PATHS = {
    "浩天": ["/new-"],
    "Morita Kagaku Kogyo": ["/news/news"],
}
REQUIRE_YEAR = frozenset({"dsm-firmenich"})
ENGLISH_ONLY = frozenset({"IFF"})

# Routing tag: a regulatory source gets the RA prompt and its own digest email.
GROUPS = {"国家卫健委": "regulatory"}

# Sites that render nothing server-side but expose a JSON feed (see json_api.py).
JSON_API = {
    "元气森林": {
        "fetch_mode": "json_api",
        "api": {
            "list_url": "https://www.yuanqisenlin.com/web/api/newsList",
            "list_items_path": "data.data",
            "list_title_key": "title",
            "list_date_key": "day",
            "list_id_key": "id",
            "list_summary_key": "short_desc",
            "detail_url": "https://www.yuanqisenlin.com/web/api/newsDetail",
            "detail_id_param": "id",
            "detail_content_path": "data.content",
            "article_url_template": "https://www.yuanqisenlin.com/newsInfo?id={id}",
        },
    },
}

_SEPARATORS_RE = re.compile(r"[;,、\n]+")


def split_keywords(raw: str) -> list[str]:
    """A keyword cell as a deduped list. Separators: ``;`` ``,`` ``、`` newline."""
    if not raw:
        return []
    keywords: list[str] = []
    seen: set[str] = set()
    for part in _SEPARATORS_RE.split(str(raw)):
        keyword = part.strip()
        if keyword and keyword.lower() not in seen:
            seen.add(keyword.lower())
            keywords.append(keyword)
    return keywords


def build_config(rows: Iterable[Sequence[Any]]) -> dict[str, Any]:
    """Pure: spreadsheet rows (header already skipped) -> the config dict."""
    sites: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    needs_url: list[dict[str, str]] = []

    for row in rows:
        name = _cell(row, 0)
        if not name:
            continue
        if name in EXCLUDED_NAMES:
            excluded.append({"name": name, "reason": "under discussion (needs web search)"})
            continue

        news_url = _cell(row, 3)
        # Every remaining row is meant to be monitored, so a missing URL is flagged
        # loudly rather than dropped.
        if not news_url.lower().startswith("http"):
            needs_url.append({"name": name, "value": news_url or "(empty)"})
            continue

        sites.append(_build_site(name, _cell(row, 1), _cell(row, 2), news_url))

    return {"sites": sites, "excluded": excluded, "needs_url": needs_url}


def _cell(row: Sequence[Any], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _build_site(name: str, keywords_raw: str, homepage: str, news_url: str) -> dict[str, Any]:
    keywords = split_keywords(keywords_raw)
    # The company's own name is always a keyword: its releases must match even when
    # the keyword cell lists only product names.
    if name.lower() not in {keyword.lower() for keyword in keywords}:
        keywords.append(name)

    site: dict[str, Any] = {
        "name": name,
        "keywords": keywords,
        "homepage": homepage,
        "news_url": news_url,
        "enabled": name not in DISABLED_NAMES,
    }
    if name in ALLOW_PATHS:
        site["allow_paths"] = ALLOW_PATHS[name]
    if name in REQUIRE_YEAR:
        site["require_year"] = True
    if name in ENGLISH_ONLY:
        site["english_only"] = True
    if name in JSON_API:
        site.update(JSON_API[name])
    if name in GROUPS:
        site["group"] = GROUPS[name]
    return site


def read_rows() -> list[Sequence[Any]]:
    """The monitoring sheet's data rows. Raises on a missing file or sheet."""
    import openpyxl

    if not LINKS_XLSX.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {LINKS_XLSX}")

    workbook = openpyxl.load_workbook(LINKS_XLSX, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(f"Sheet '{SHEET_NAME}' not found. Sheets: {workbook.sheetnames}")

    rows = list(workbook[SHEET_NAME].iter_rows(values_only=True))
    return rows[1:]  # skip the header


def main() -> int:
    configure_console()
    try:
        rows = read_rows()
    except (FileNotFoundError, KeyError) as exc:
        logger.error("❌ %s", exc)
        return 1

    config = build_config(rows)
    SITES_PATH.parent.mkdir(parents=True, exist_ok=True)
    SITES_PATH.write_text(
        json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    logger.info("✅ Wrote %s", SITES_PATH)
    logger.info("   Monitored sites: %d", len(config["sites"]))
    logger.info("   Excluded (under discussion): %d", len(config["excluded"]))
    if config["needs_url"]:
        logger.warning("   ⚠️  Rows still missing a valid news URL: %d", len(config["needs_url"]))
        for item in config["needs_url"]:
            logger.warning("       - %s  (current value: %s)", item["name"], item["value"])
    return 0


if __name__ == "__main__":
    sys.exit(main())
